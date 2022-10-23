from openpyxl import Workbook
from scheme import Goods, Positions, Orders, Barcodes, Base
from run import session


wb = Workbook()
ws = wb.active

ws['A1'] = 'SKU'
ws['B1'] = 'Артикул'
ws['C1'] = 'Бренд'
ws['D1'] = 'Категория'
ws['E1'] = 'Цена'
ws['F1'] = 'Цена со скидкой'

for i, good in enumerate(session.query(Goods).all()):
    ws[f'A{i+2}'] = good.sku
    ws[f'B{i+2}'] = good.article
    ws[f'C{i+2}'] = good.brand
    ws[f'D{i+2}'] = good.category
    ws[f'E{i+2}'] = good.price
    ws[f'F{i+2}'] = good.price_after_spp

ws = wb.create_sheet('Позиции')

ws['A1'] = 'Дата'
ws['B1'] = 'SKU'
ws['D1'] = 'Позиция'
ws['C1'] = 'Запрос'

for i, position in enumerate(session.query(Positions).all()):
    ws[f'A{i+2}'] = position.date
    ws[f'B{i+2}'] = position.sku
    ws[f'C{i+2}'] = position.query
    ws[f'D{i+2}'] = position.position

ws = wb.create_sheet('Заказы')

ws['B1'] = 'SKU'
ws['A1'] = 'Дата'
ws['C1'] = 'Баркод'
ws['D1'] = 'FBS'
ws['E1'] = 'FBO'

for i, order in enumerate(session.query(Orders).all()):
    ws[f'A{i+2}'] = order.sku
    ws[f'B{i+2}'] = order.date
    ws[f'C{i+2}'] = order.barcode
    ws[f'D{i+2}'] = order.fbs_count
    ws[f'E{i+2}'] = order.fbo_count

wb.save('dump.xlsx')